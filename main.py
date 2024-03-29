from app.device_compile import cisco_compile, palo_compile, fortigate_compile
from app.mail import send_mail
from datetime import date, timedelta, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side
import re
import xml.etree.ElementTree as ET
import os
import logging
import sys
import pathlib


def main():
    
    path = pathlib.Path(__file__).parent.resolve()
    
    # initiate log file
    timestamp = f"{datetime.now().month}-{datetime.now().day}-{datetime.now().year}_{datetime.now().hour}:{datetime.now().minute}"
    logging.basicConfig(filename=f"{path}/logs/{timestamp}.log", level=logging.INFO)
    
    # create devices, log errors to log file
    try:
        palo_list = palo_compile()
    except Exception as e:
        logging.error(f"Could not compile paloalto.ini: {e}")
        sys.exit()
    
    try:
        cisco_list = cisco_compile()
    except Exception as e:
        logging.error(f"Could not compile cisco.ini: {e}")
        sys.exit()
        
    try:
        fortigate_list = fortigate_compile()
    except Exception as e:
        logging.error(f"Could not compile fortigate.ini: {e}")
        sys.exit()

    # set command to send to cisco devices
    palo_xpath = "<show><high-availability><state></state></high-availability></show>"
    cisco_command = "show standby brief"

    # run commands on devices and create dict of responses
    palo_data = {}
    for dev in palo_list:
        try:
            palo_data[dev.name] = ET.fromstring(dev.send_xpath(palo_xpath).content).find(
            'result').find('group').find('local-info').find('state').text
        except Exception as e:
            logging.error(f"Invalid response from Palo Alto device {dev.name}: {e}")
            palo_data[dev.name] = "Invalid Response"
    
    cisco_data = {}
    for dev in cisco_list:
        try:
            cisco_data[dev.name] = re.findall(r"\A[^,]+?P (Active|Standby)", str(dev.send_cmd(cisco_command)))
        except Exception as e:
            logging.error(f"Invalid response from Cisco device {dev.name}: {e}")
            cisco_data[dev.name] = "Invalid Response"
            
            
    fortigate_data = {}
    for dev in fortigate_list:
        try:
            active = dev.get_hostname()
            if dev.primary == active:
                fortigate_data[dev.primary] = "Active"
                fortigate_data[dev.secondary] = "Passive"
            elif dev.secondary == active:
                fortigate_data[dev.primary] = "Passive"
                fortigate_data[dev.secondary] = "Active"
            else:
                fortigate_data[dev.primary] = "Invalid Response"
                fortigate_data[dev.secondary] = "Invalid Response"
        except Exception as e:
            logging.error(f"Invalid response from Fortigate device {dev.name}: {e}")
            fortigate_data[dev.primary] = "Invalid Response"
            fortigate_data[dev.secondary] = "Invalid Response"
                


    # user responses to create spreadsheet report
    try:
        report = get_report(palo_data, cisco_data, fortigate_data)
    except Exception as e:
        logging.error(f"Could not generate report: {e}")
        sys.exit()
    
    # compare results to previous day to see if anything changed
    try:
        changed = compare(palo_data, cisco_data, fortigate_data)
    except Exception as e:
        logging.error(f"Could not compare today's data to historical data: {e}")
        
    # generate the text for the email message
    if len(changed) == 0:
        text = "No devices have changed state from the previous day"
    else:
        text = "The following devices have changed state:\n\n" + ', '.join(changed)
    
    # log result to log file as well
    logging.info(text)

    # send report via mail
    try:
        send_mail(text, report)
    except Exception as e:
        logging.error(f"Could not send mail: {e}")

    # log completion to log file if successful
    logging.info(f"Completed running at {datetime.now()}")


def get_report(palo, cisco, fortigate):
    # open new excel sheet and add headers
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sitrep"
        ws.append(['Device Name', 'State'])
        ws.append(['Cisco Devices'])
    except Exception as e:
        raise RuntimeError(f"Could not initialize workbook: {e}")

    # add all devices and their states to spreadsheet
    try:
        for name, state in cisco.items():
            ws.append([name, state[0]])
    except Exception as e:
       raise RuntimeError(f"Could not add Cisco data to workbook: {e}")
   
    try:
        ws.append(['Palo Alto Devices'])
        for name, state in palo.items():
            ws.append([name, state.title()])
    except Exception as e:
       raise RuntimeError(f"Could not add Palo Alto data to workbook: {e}")
   
    try:
        ws.append(['Fortigate Devices'])
        for name, state in fortigate.items():
            ws.append([name, state.title()])
    except Exception as e:
       raise RuntimeError(f"Could not add Fortigate data to workbook: {e}")
   
    try:
        thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
        c = ws['A1']
        c.style = '40 % - Accent1'
        c.font = Font(bold=True)
        c.border = thin_border
        c = ws['B1']
        c.style = '40 % - Accent1'
        c.font = Font(bold=True)
        c.border = thin_border
        for row in ws.iter_rows():
            if row[0].value == 'Cisco Devices' or row[0].value == 'Palo Alto Devices' or row[0].value == 'Fortigate Devices':
                row[0].style = '40 % - Accent2'
                row[0].font = Font(bold=True)
                row[0].border = thin_border
                row[1].style = '40 % - Accent2'
                row[1].font = Font(bold=True)
                row[1].border = thin_border
            elif row[0].value == 'Device Name':
                pass
            elif row[0].value:
                row[0].style = '20 % - Accent3'
                row[0].border = thin_border
                row[1].style = '20 % - Accent3'
                row[1].border = thin_border
                
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 10
                
    except:
        pass
                

    # save the excel sheet with todays date to the /reports directory
    try:
        path = pathlib.Path(__file__).parent.resolve()
        report = f"{path}/reports/{str(date.today())}-sitrep.xlsx"
        wb.save(report)
    except Exception as e:
       raise RuntimeError(f"Could not save workbook: {e}")
   
    # return the report file path
    return report


def compare(palo, cisco, fortigate):
    # get yesterday's date
    today = date.today()
    yesterday = today - timedelta(days=1)

    # initiate list of changed devices
    changed = []

    # try to open the report from the previous day
    try:
        path = pathlib.Path(__file__).parent.resolve()
        wb = load_workbook(f"{path}/reports/{str(yesterday)}-sitrep.xlsx")
        ws = wb["Sitrep"]
    except Exception as e:
        logging.error(f"Could not open previous report: {e}") 
        return ["Yesterday\'s sitrep not detected, see log file for details"]

    # compare each device state to the previous day's state
    try:
        for row in range(1, (len(palo)+len(cisco)+len(fortigate)+len(fortigate)+3)):
            dev = ws[f"A{row}"].value
            old_state = ws[f"B{row}"].value
            if dev is None or old_state is None:
                continue
            if dev in cisco:
                state = cisco[dev][0]
            elif dev in palo:
                state = palo[dev]
            elif dev in fortigate:
                state = fortigate[dev]
            else:
                continue
            if state.lower() != old_state.lower():
                changed.append(dev)
    except Exception as e:
        logging.error(f"Could not parse previous report: {e}") 
        return ["Could not parse yesterday\'s sitrep, see log file for details"]

    # return the list of devices that changed state
    return changed


if __name__ == "__main__":
    main()
