import devices
import smtplib
from datetime import date, timedelta
from openpyxl import Workbook, load_workbook
import re
from xapi import PanXapiError, PanXapi
import xml.etree.ElementTree as ET
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from config import api


def main():
    # create devices
    palo_list = [
        PanXapi(tag='DC1_VPN', api_key=api, hostname='172.16.50.50'),
        PanXapi(tag='601_VPN', api_key=api, hostname='172.16.50.51')
    ]

    cisco_list = [
        devices.create("DC1-6807", "172.16.50.60", "cisco_ios",
                       "lab", "/Users/msexton/.ssh/lab"),
        devices.create("601-6807", "172.16.50.60", "cisco_ios",
                       "lab", "/Users/msexton/.ssh/lab")
    ]

    p_command = "<show><high-availability><state></state></high-availability></show>"
    c_command = "show standby brief"

    # run uptime on devices and create dict of responses
    palo_data = {dev.tag: ET.fromstring(dev.op(p_command).read()).find(
        'result').find('group').find('local-info').find('state').text for dev in palo_list}

    cisco_data = {dev.name: re.findall(
        r"\A[^,]+?P (Active|Standby)", str(dev.send_cmd(c_command))) for dev in cisco_list}
    get_report(palo_data, cisco_data)
    changed = compare(palo_data, cisco_data)

    if len(changed) == 0:
        text = "No changes"
    else:
        text = "The following devices have changed state:\n\n" + \
            '\n'.join(changed)
    print(text)
    # send_mail(f"reports/{str(date.today())}-sitrep.xlsx", text)


def get_report(palo, cisco):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sitrep"
    ws.append(['Device Name', 'State'])

    for name, state in cisco.items():
        ws.append([name, state[0]])

    for name, state in palo.items():
        ws.append([name, state.title()])

    wb.save(f"reports/{str(date.today())}-sitrep.xlsx")


def compare(palo, cisco):
    today = date.today()
    yesterday = today - timedelta(days=1)

    changed = []

    try:
        wb = load_workbook(f"reports/{str(yesterday)}-sitrep.xlsx")
        ws = wb["Sitrep"]
    except:
        return ['Yesterday\'s sitrep not detected']

    for row in range(1, 10):
        dev = ws[f"A{row}"].value
        old_state = ws[f"B{row}"].value

        if dev is None or old_state is None:
            continue

        if dev in cisco:
            state = cisco[dev][0]
        elif dev in palo:
            state = palo[dev]
        else:
            continue

        if state != old_state:
            changed.append(dev)

    return changed


def send_mail(report, text):
    SERVER = 'smtp.novanthealth.org'
    PORT = 443

    msg = MIMEMultipart()
    body_part = MIMEText(text, 'plain')
    msg['Subject'] = f"Sitrep {date.today()}"
    msg['From'] = 'nse_sitrep@novanthealth.org'
    msg['To'] = 'cpsnse@novanthealth.org'
    # Add body to email
    msg.attach(body_part)
    # open and read the CSV file in binary
    with open(report, 'rb') as file:
        # Attach the file with filename to the email
        msg.attach(MIMEApplication(file.read(), Name=report))

    # Create SMTP object
    smtp_obj = smtplib.SMTP(SERVER, PORT)

    # Convert the message to a string and send it
    smtp_obj.sendmail(msg['From'], msg['To'], msg.as_string())
    smtp_obj.quit()


if __name__ == "__main__":
    main()
